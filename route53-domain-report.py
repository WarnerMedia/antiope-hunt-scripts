#!/usr/bin/env python3
import boto3
import botocore
from boto3.dynamodb.conditions import Key

import re
import requests
from requests_aws4auth import AWS4Auth
from elasticsearch import Elasticsearch, RequestsHttpConnection

import json
import datetime
import yaml
import zipfile

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
except:
    print("Must Install the openpyxl modules")
    exit(-1)

try:
    from antiope.aws_account import *
except:
    print("Must install the Antiope python modules")
    exit(-1)


import logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)
logging.getLogger('botocore').setLevel(logging.WARNING)
logging.getLogger('boto3').setLevel(logging.WARNING)
logging.getLogger('urllib3').setLevel(logging.WARNING)
logging.getLogger('elasticsearch').setLevel(logging.WARNING)


def main(args, logger):
    domain_rows = []
    yaml_datas = []
    yaml_names = []

    antiope_config = AntiopeConfig(SSMParam=args.ssm_param)

    client = boto3.client('ssm')
    response = client.get_parameter(Name=args.es_param)
    es_config = json.loads(response['Parameter']['Value'])

    for domain in get_domain_contents(es_config['ClusterEndpoint'], os.environ['AWS_DEFAULT_REGION']):
        logger.info(f"Domain: {domain['resourceName']}")

        name = domain['resourceName']
        creation_date = domain['resourceCreationTime']
        creation_date = creation_date.split(' ')[0]

        account = AWSAccount(domain['awsAccountId'], config=antiope_config)

        nameservers = ""
        for ns in domain['configuration']['Nameservers']:
            nameservers += f"{ns['Name']}\n"

        row_to_add = [
            name,
            creation_date,
            account.account_name,
            account.exec_sponsor_email,
            account.payer_name,
            yaml.dump(domain['configuration']['RegistrantContact']),
            yaml.dump(domain['configuration']['AdminContact']),
            yaml.dump(domain['configuration']['TechContact']),
            nameservers,
            domain['supplementaryConfiguration']['TransferLock'],
            domain['configuration']['ExpirationDate'],
            domain['configuration']['AutoRenew'],
            domain['configuration']['RegistrarName'],
        ]

        logger.debug(f"{row_to_add}")
        domain_rows.append(row_to_add)

    wb = create_workbook_from_domains(domain_rows)
    wb.save(args.filename)


def create_workbook_from_domains(domain_rows):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Creation Date', 'Account Name', 'AWS Account Owner', 'AWS Account Payer', 'RegistrantContact', 'AdminContact', 'TechContact', 'Name Servers', 'TransferLock', 'ExpirationDate', 'AutoRenew', 'RegistrarName'])
    for domain in domain_rows[::-1]:
        ws.append(domain)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 50
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['M'].width = 17

    ws.auto_filter.ref = 'A1:{}{:d}'.format(get_column_letter(ws.max_column), ws.max_row)

    for cell in ws["1:1"]:
        cell.font = Font(bold=True)

    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.style.alignment.wrap_text=True

    return wb


def get_domain_contents(es_endpoint, region):
    output = []

    service = 'es'
    credentials = boto3.Session().get_credentials()
    awsauth = AWS4Auth(credentials.access_key, credentials.secret_key, region, service, session_token=credentials.token)

    es = Elasticsearch(
        hosts=[{'host': es_endpoint, 'port': 443}],
        http_auth=awsauth,
        use_ssl=True,
        verify_certs=True,
        connection_class=RequestsHttpConnection
    )
    logger.debug(es.info())

    index_name = "resources_route53_domain"

    query = {
        "bool": {
            "must": [
                {"range": {"configurationItemCaptureTime": {"gte": "now-3h"}}}
            ]
        }
    }

    res = es.search(index=index_name, size=10000, body={"query": query})
    logger.debug(f"found {len(res['hits']['hits'])} trail records")

    for hit in res['hits']['hits']:
        logger.debug(json.dumps(hit, sort_keys=True, default=str, indent=2))
        doc = hit['_source']
        output.append(doc)

    return(output)


def do_args():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", help="print debugging info", action='store_true')
    parser.add_argument("--ssm-param", help="ssm parameter with Antiope configs", default="antiope-aws")
    parser.add_argument("--es-param", help="ssm parameter with ElasticSearch configs", default="antiope-aws")
    parser.add_argument("--filename", help="Base filename for the output files", default=f"aws-domains-{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx")
    args = parser.parse_args()
    return(args)


if __name__ == '__main__':
    args = do_args()

    # Logging idea stolen from: https://docs.python.org/3/howto/logging.html#configuring-logging
    # create console handler and set level to debug
    ch = logging.StreamHandler()
    if args.debug:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)
    # create formatter
    # formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    formatter = logging.Formatter('%(levelname)s - %(message)s')
    # add formatter to ch
    ch.setFormatter(formatter)
    # add ch to logger
    logger.addHandler(ch)

    # Wrap in a handler for Ctrl-C
    try:
        main(args, logger)
    except KeyboardInterrupt:
        exit(1)


